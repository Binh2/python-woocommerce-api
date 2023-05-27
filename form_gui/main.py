from typing import List
from openpyxl import *
from tkinter import *
from string import ascii_uppercase
from App import App

OVERFLOW = 100

filename = './excel.xlsx'
wb = load_workbook(filename)

sheet = wb.active

# columns = [ 'A'...'Z', 'AA'...'AZ', 'BA'...'BZ', ..., 'ZZ' ]
columns = [ letter for letter in ascii_uppercase ] + [ left + right for left in ascii_uppercase for right in ascii_uppercase ]
column_names = [ "Name", "Price", "Images", "Weight(kg)", "Length(cm)", "Width(cm)", "Height(cm)", "Attribute 1 name", "Attribute 1 value(s)", "Attribute 1 visible", "Attribute 1 global", "Attribute 1 default", "Attribute 2 name", "Attribute 2 value(s)", "Attribute 2 visible", "Attribute 2 global", "Attribute 2 default", "Attribute 3 name", "Attribute 3 value(s)", "Attribute 3 visible", "Attribute 3 global", "Attribute 3 default", "Attribute 4 name", "Attribute 4 value(s)", "Attribute 4 visible", "Attribute 4 global", "Attribute 4 default", "Attribute 5 name", "Attribute 5 value(s)", "Attribute 5 visible", "Attribute 5 global", "Attribute 5 default", "Attribute 6 name", "Attribute 6 value(s)", "Attribute 6 visible", "Attribute 6 global", "Attribute 6 default", "Attribute 7 name", "Attribute 7 value(s)", "Attribute 7 visible", "Attribute 7 global", "Attribute 7 default", "Attribute 8 name", "Attribute 8 value(s)", "Attribute 8 visible", "Attribute 8 global", "Attribute 8 default" ]


def set_column_names(sheet):
  '''Change columns's name and columns's size'''
  for index, column, column_name in zip(range(1, 10000), columns, column_names):
    sheet.cell(row=1, column=index).value = column_name
    sheet.column_dimensions[column].width = len(column_name)

def insert(sheet, fields: List[Entry]):
  if all([ field.get() for field in root.fields ]):
    print("Empty form")
    return
  
  current_row = sheet.max_row
  current_column = sheet.max_column

  for index, field in zip(range(1, 10000), fields):
    sheet.cell(row=current_row + 1, column=index).value = field.get()

  wb.save(filename)
  fields[0].focus_set()
  root.clear_form()






if __name__ == "__main__":
  set_column_names(sheet)
  
  root = App(Tk(), column_names, lambda fields: insert(sheet, fields))
  root.tk.mainloop()
